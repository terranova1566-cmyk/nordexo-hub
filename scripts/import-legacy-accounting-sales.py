#!/usr/bin/env python3
"""Normalize accounting export workbooks and produce CSV for legacy_sales_data imports."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
from collections import Counter
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import CALENDAR_WINDOWS_1900, from_excel


FIELD_ALIASES = {
    "marketplace": ["Marknadsplats", "Platform", "Marketplace"],
    "sku": ["Egen referens", "SKU"],
    "sold_date": ["Skapad", "Date", "Date sold"],
    "total_price": ["Total inkl moms och frakt", "Price", "Total price"],
    "currency": ["Valuta", "Currency"],
    "amount_sold": ["Antal", "Amount"],
}

PLATFORM_CODE_MAP = {
    "CDON DK": "CD-DK",
    "CDON FI": "CD-FI",
    "CDON NO": "CD-NO",
    "CDON SE": "CD-SE",
    "FYNDIQ SE": "FQ-SE",
    "LETSDEAL NO": "LD-NO",
    "LETSDEAL SE": "LD-SE",
}

NUM_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def to_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def resolve_columns(header_values: list[object], file_name: str) -> dict[str, int]:
    by_header = {normalize_header(h): i for i, h in enumerate(header_values)}
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for field, aliases in FIELD_ALIASES.items():
        idx = None
        for alias in aliases:
            key = normalize_header(alias)
            if key in by_header:
                idx = by_header[key]
                break
        if idx is None:
            missing.append(f"{field} ({' / '.join(aliases)})")
        else:
            resolved[field] = idx

    if missing:
        raise SystemExit(
            f"Missing required columns in {file_name}: {', '.join(missing)}"
        )
    return resolved


def parse_decimal(value: object) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        if value.is_nan() or value.is_infinite():
            return None
        return value
    if isinstance(value, int):
        parsed = Decimal(value)
        if parsed.is_nan() or parsed.is_infinite():
            return None
        return parsed
    if isinstance(value, float):
        if value != value:
            return None
        parsed = Decimal(str(value))
        if parsed.is_nan() or parsed.is_infinite():
            return None
        return parsed

    text = to_text(value)
    if not text:
        return None
    compact = text.replace(" ", "").replace("\u00a0", "")
    try:
        parsed = Decimal(compact.replace(",", "."))
        if parsed.is_nan() or parsed.is_infinite():
            return None
        return parsed
    except Exception:
        match = NUM_RE.search(compact)
        if not match:
            return None
        try:
            parsed = Decimal(match.group(0).replace(",", "."))
            if parsed.is_nan() or parsed.is_infinite():
                return None
            return parsed
        except Exception:
            return None


def parse_date(value: object) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        parsed = value.date()
    elif isinstance(value, dt.date):
        parsed = value
    elif isinstance(value, (int, float, Decimal)):
        try:
            converted = from_excel(value, CALENDAR_WINDOWS_1900)
        except Exception:
            return None
        if isinstance(converted, dt.datetime):
            parsed = converted.date()
        elif isinstance(converted, dt.date):
            parsed = converted
        else:
            return None
    else:
        text = to_text(value)
        if not text:
            return None
        parsed = None
        for fmt in (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%d/%m/%Y",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y %H:%M:%S",
        ):
            try:
                parsed = dt.datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        if parsed is None:
            try:
                parsed = dt.datetime.fromisoformat(text.replace("Z", "+00:00")).date()
            except ValueError:
                return None

    if parsed.year < 2000 or parsed.year > 2100:
        return None
    return parsed


def split_marketplace(raw: Optional[str]) -> tuple[Optional[str], Optional[str], Optional[str]]:
    if not raw:
        return None, None, None
    text = re.sub(r"\s+", " ", raw.strip())
    if not text:
        return None, None, None

    tokens = text.split(" ")
    if len(tokens) % 2 == 0:
        half = len(tokens) // 2
        left = " ".join(tokens[:half]).strip()
        right = " ".join(tokens[half:]).strip()
        if left and right and left.lower() == right.lower():
            return left, right, text

    return text, text, text


def normalize_platform_code(code: Optional[str]) -> Optional[str]:
    if not code:
        return code
    mapped = PLATFORM_CODE_MAP.get(code.upper())
    return mapped or code


def decimal_to_str(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return str(normalized.quantize(Decimal("1")))
    return format(normalized, "f")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import accounting legacy sales workbooks")
    parser.add_argument("inputs", nargs="+", help="Input accounting .xlsx files.")
    parser.add_argument("--batch-label", required=True, help="Batch label for inserted rows.")
    parser.add_argument("--output-csv", required=True, help="Merged normalized CSV path.")
    parser.add_argument("--output-report", required=True, help="JSON report path.")
    parser.add_argument("--output-platforms-csv", required=True, help="Unique platform code/name CSV path.")
    parser.add_argument(
        "--output-load-sql",
        required=False,
        help="Optional SQL file with psql \\copy command for legacy_sales_data.",
    )
    return parser.parse_args()


def write_load_sql(path: Path, csv_path: Path) -> None:
    escaped_csv = str(csv_path).replace("'", "''")
    sql = (
        "\\copy public.legacy_sales_data("
        "batch_label,source_file,source_row_number,seller_platform,sku,sku_norm,sold_date,amount_sold,total_price,currency"
        f") from '{escaped_csv}' with (format csv, header true);\n"
    )
    path.write_text(sql, encoding="utf-8")


def main() -> None:
    args = parse_args()
    input_files = [Path(p).expanduser().resolve() for p in args.inputs]
    out_csv = Path(args.output_csv).expanduser().resolve()
    out_report = Path(args.output_report).expanduser().resolve()
    out_platforms = Path(args.output_platforms_csv).expanduser().resolve()
    out_load_sql = Path(args.output_load_sql).expanduser().resolve() if args.output_load_sql else None

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    out_report.parent.mkdir(parents=True, exist_ok=True)
    out_platforms.parent.mkdir(parents=True, exist_ok=True)
    if out_load_sql:
        out_load_sql.parent.mkdir(parents=True, exist_ok=True)

    totals = Counter()
    files_report: list[dict[str, object]] = []
    platform_counts: Counter[tuple[str, str]] = Counter()

    min_date: Optional[str] = None
    max_date: Optional[str] = None

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "batch_label",
                "source_file",
                "source_row_number",
                "seller_platform",
                "sku",
                "sku_norm",
                "sold_date",
                "amount_sold",
                "total_price",
                "currency",
            ]
        )

        for input_file in input_files:
            wb = load_workbook(input_file, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]

            header_values = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            col_index = resolve_columns(header_values, input_file.name)

            file_stats = Counter()
            file_stats["file"] = str(input_file)
            file_stats["worksheet"] = ws.title
            file_stats["header_row_skipped"] = True

            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                file_stats["rows_scanned"] += 1
                totals["rows_scanned"] += 1

                marketplace_raw = to_text(row[col_index["marketplace"]])
                sku = to_text(row[col_index["sku"]])
                sold_date = parse_date(row[col_index["sold_date"]])
                total_price = parse_decimal(row[col_index["total_price"]])
                currency = to_text(row[col_index["currency"]])
                amount_sold = parse_decimal(row[col_index["amount_sold"]])

                if amount_sold is None:
                    file_stats["dropped_missing_amount"] += 1
                    totals["dropped_missing_amount"] += 1
                    continue
                if not sku:
                    file_stats["dropped_missing_sku"] += 1
                    totals["dropped_missing_sku"] += 1
                    continue
                if sold_date is None:
                    file_stats["dropped_missing_date"] += 1
                    totals["dropped_missing_date"] += 1
                    continue
                if total_price is None:
                    file_stats["dropped_missing_total_price"] += 1
                    totals["dropped_missing_total_price"] += 1
                    continue
                if not currency:
                    file_stats["dropped_missing_currency"] += 1
                    totals["dropped_missing_currency"] += 1
                    continue

                marketplace_code, marketplace_name, marketplace_source = split_marketplace(marketplace_raw)
                seller_platform = normalize_platform_code(marketplace_code) or ""
                platform_name = marketplace_name or ""
                if seller_platform or platform_name:
                    platform_counts[(seller_platform, platform_name)] += 1

                sold_date_text = sold_date.isoformat()
                if min_date is None or sold_date_text < min_date:
                    min_date = sold_date_text
                if max_date is None or sold_date_text > max_date:
                    max_date = sold_date_text

                writer.writerow(
                    [
                        args.batch_label,
                        input_file.name,
                        row_number,
                        seller_platform,
                        sku,
                        re.sub(r"\s+", "", sku).upper(),
                        sold_date_text,
                        decimal_to_str(amount_sold),
                        decimal_to_str(total_price),
                        currency,
                    ]
                )

                if marketplace_source:
                    file_stats["marketplace_source_distinct"] = file_stats.get(
                        "marketplace_source_distinct", 0
                    )
                file_stats["rows_written"] += 1
                totals["rows_written"] += 1

            files_report.append(
                {
                    "file": file_stats.get("file"),
                    "worksheet": file_stats.get("worksheet"),
                    "rows_scanned": int(file_stats.get("rows_scanned", 0)),
                    "rows_written": int(file_stats.get("rows_written", 0)),
                    "dropped_missing_amount": int(file_stats.get("dropped_missing_amount", 0)),
                    "dropped_missing_sku": int(file_stats.get("dropped_missing_sku", 0)),
                    "dropped_missing_date": int(file_stats.get("dropped_missing_date", 0)),
                    "dropped_missing_total_price": int(file_stats.get("dropped_missing_total_price", 0)),
                    "dropped_missing_currency": int(file_stats.get("dropped_missing_currency", 0)),
                    "header_row_skipped": True,
                }
            )

    with out_platforms.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["platform_code", "platform_name", "rows"])
        for (platform_code, platform_name), count in sorted(platform_counts.items()):
            writer.writerow([platform_code, platform_name, count])

    report = {
        "batch_label": args.batch_label,
        "generated_at_utc": dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "inputs": [str(p) for p in input_files],
        "output_csv": str(out_csv),
        "output_platforms_csv": str(out_platforms),
        "date_range": {"min_sold_date": min_date, "max_sold_date": max_date},
        "totals": {
            "rows_scanned": int(totals.get("rows_scanned", 0)),
            "rows_written": int(totals.get("rows_written", 0)),
            "dropped_missing_amount": int(totals.get("dropped_missing_amount", 0)),
            "dropped_missing_sku": int(totals.get("dropped_missing_sku", 0)),
            "dropped_missing_date": int(totals.get("dropped_missing_date", 0)),
            "dropped_missing_total_price": int(totals.get("dropped_missing_total_price", 0)),
            "dropped_missing_currency": int(totals.get("dropped_missing_currency", 0)),
        },
        "unique_platforms": len(platform_counts),
        "files": files_report,
    }
    out_report.write_text(json.dumps(report, indent=2), encoding="utf-8")

    if out_load_sql:
        write_load_sql(out_load_sql, out_csv)


if __name__ == "__main__":
    main()
