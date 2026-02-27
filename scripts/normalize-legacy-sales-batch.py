#!/usr/bin/env python3
"""Normalize legacy sales Excel files into one clean CSV batch."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
from decimal import Decimal
from itertools import chain
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import CALENDAR_WINDOWS_1900, from_excel


GENERIC_HEADERS = ["column1", "column2", "column3", "column4", "column5"]
NUMERIC_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def to_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def looks_like_header(first_row: list[object]) -> bool:
    normalized = [to_text(v).lower() if to_text(v) else "" for v in first_row[:5]]
    if normalized == GENERIC_HEADERS:
        return True
    return (
        "sku" in normalized
        or "amount" in normalized
        or "date" in normalized
        or "sold_date" in normalized
    )


def normalize_sku(raw_sku: object) -> tuple[Optional[str], Optional[str]]:
    sku = to_text(raw_sku)
    if not sku:
        return None, None
    sku_norm = re.sub(r"\s+", "", sku).upper()
    return sku, sku_norm


def parse_amount(value: object) -> Optional[Decimal]:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, int):
        return Decimal(value)

    if isinstance(value, float):
        if value != value:  # NaN check
            return None
        return Decimal(str(value))

    text = to_text(value)
    if not text:
        return None

    match = NUMERIC_RE.search(text)
    if not match:
        return None

    numeric = match.group(0).replace(",", ".")
    try:
        return Decimal(numeric)
    except Exception:
        return None


def parse_date_string(text: str) -> Optional[dt.date]:
    text = text.strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            parsed = dt.datetime.strptime(text, fmt)
            return parsed.date()
        except ValueError:
            continue

    # Parse plain ISO datetime strings like 2021-12-01T12:49:00
    try:
        parsed = dt.datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.date()
    except ValueError:
        return None


def is_reasonable_sales_date(value: dt.date) -> bool:
    return 2000 <= value.year <= 2100


def parse_sold_date(cell_value: object) -> Optional[dt.date]:
    if cell_value is None:
        return None

    if isinstance(cell_value, dt.datetime):
        parsed = cell_value.date()
        return parsed if is_reasonable_sales_date(parsed) else None

    if isinstance(cell_value, dt.date):
        return cell_value if is_reasonable_sales_date(cell_value) else None

    if isinstance(cell_value, dt.time):
        return None

    if isinstance(cell_value, (int, float, Decimal)):
        try:
            parsed = from_excel(cell_value, CALENDAR_WINDOWS_1900)
        except Exception:
            return None
        if isinstance(parsed, dt.datetime):
            parsed_date = parsed.date()
        elif isinstance(parsed, dt.date):
            parsed_date = parsed
        else:
            return None
        return parsed_date if is_reasonable_sales_date(parsed_date) else None

    text = to_text(cell_value)
    if not text:
        return None
    parsed = parse_date_string(text)
    if not parsed:
        return None
    return parsed if is_reasonable_sales_date(parsed) else None


def decimal_to_string(value: Decimal) -> str:
    normalized = value.normalize()
    # Keep integers without trailing .0
    if normalized == normalized.to_integral():
        return str(normalized.quantize(Decimal("1")))
    return format(normalized, "f")


def write_load_sql(sql_path: Path, table_name: str, csv_path: Path) -> None:
    escaped_csv = str(csv_path).replace("'", "''")
    sql = (
        f"\\copy {table_name}(batch_label,source_file,source_row_number,seller_platform,sku,sku_norm,sold_date,amount_sold) "
        f"from '{escaped_csv}' with (format csv, header true);\n"
    )
    sql_path.write_text(sql, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize messy legacy sales Excel files into a clean merged CSV."
    )
    parser.add_argument(
        "inputs",
        nargs="+",
        help="Input .xlsx files to merge.",
    )
    parser.add_argument(
        "--batch-label",
        required=True,
        help="Batch label written to each row (for example: legacy_batch_1).",
    )
    parser.add_argument(
        "--output-csv",
        required=True,
        help="Path for merged normalized CSV output.",
    )
    parser.add_argument(
        "--output-report",
        required=True,
        help="Path for JSON report with row counts and dropped-row reasons.",
    )
    parser.add_argument(
        "--output-load-sql",
        help="Optional path for a generated psql \\copy SQL file.",
    )
    parser.add_argument(
        "--table-name",
        default="public.legacy_sales_data",
        help="Destination table used in generated load SQL.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    input_paths = [Path(p).expanduser().resolve() for p in args.inputs]
    output_csv = Path(args.output_csv).expanduser().resolve()
    output_report = Path(args.output_report).expanduser().resolve()
    output_load_sql = Path(args.output_load_sql).expanduser().resolve() if args.output_load_sql else None

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    output_report.parent.mkdir(parents=True, exist_ok=True)
    if output_load_sql:
        output_load_sql.parent.mkdir(parents=True, exist_ok=True)

    totals = {
        "rows_scanned": 0,
        "rows_written": 0,
        "dropped_missing_amount": 0,
        "dropped_missing_sku": 0,
        "dropped_missing_date": 0,
    }

    report_files: list[dict[str, object]] = []
    min_date: Optional[str] = None
    max_date: Optional[str] = None

    with output_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "batch_label",
                "source_file",
                "source_row_number",
                "seller_platform",
                "sku",
                "sku_norm",
                "sold_date",
                "amount_sold",
            ]
        )

        for input_path in input_paths:
            wb = load_workbook(input_path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]

            file_stats = {
                "file": str(input_path),
                "worksheet": ws.title,
                "rows_scanned": 0,
                "rows_written": 0,
                "dropped_missing_amount": 0,
                "dropped_missing_sku": 0,
                "dropped_missing_date": 0,
                "header_row_skipped": False,
            }

            row_iter: Iterable[tuple[object, ...]] = ws.iter_rows(min_col=1, max_col=5, values_only=True)
            first_row = next(row_iter, None)
            if first_row is None:
                report_files.append(file_stats)
                continue

            first_row_values = list(first_row)
            start_row_number = 1
            if looks_like_header(first_row_values):
                file_stats["header_row_skipped"] = True
                start_row_number = 2
            else:
                # Process the first row as data.
                row_iter = chain([tuple(first_row_values)], row_iter)

            for row_index, row in enumerate(row_iter, start=start_row_number):
                seller_platform = to_text(row[0] if len(row) > 0 else None)
                amount = parse_amount(row[1] if len(row) > 1 else None)
                sku, sku_norm = normalize_sku(row[2] if len(row) > 2 else None)
                sold_date = parse_sold_date(row[4] if len(row) > 4 else None)

                file_stats["rows_scanned"] += 1

                if amount is None:
                    file_stats["dropped_missing_amount"] += 1
                    continue
                if sku is None:
                    file_stats["dropped_missing_sku"] += 1
                    continue
                if sold_date is None:
                    file_stats["dropped_missing_date"] += 1
                    continue

                sold_date_text = sold_date.isoformat()
                if min_date is None or sold_date_text < min_date:
                    min_date = sold_date_text
                if max_date is None or sold_date_text > max_date:
                    max_date = sold_date_text

                writer.writerow(
                    [
                        args.batch_label,
                        input_path.name,
                        row_index,
                        seller_platform or "",
                        sku,
                        sku_norm or "",
                        sold_date_text,
                        decimal_to_string(amount),
                    ]
                )
                file_stats["rows_written"] += 1

            for key in totals:
                if key in file_stats:
                    totals[key] += int(file_stats[key])  # type: ignore[arg-type]

            report_files.append(file_stats)

    report = {
        "batch_label": args.batch_label,
        "generated_at_utc": dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "inputs": [str(p) for p in input_paths],
        "output_csv": str(output_csv),
        "date_range": {"min_sold_date": min_date, "max_sold_date": max_date},
        "totals": totals,
        "files": report_files,
    }
    output_report.write_text(json.dumps(report, indent=2), encoding="utf-8")

    if output_load_sql:
        write_load_sql(output_load_sql, args.table_name, output_csv)


if __name__ == "__main__":
    main()
